"""
把 sampledata/商务数据/ 下的 .xlsx 文件转换为 .txt 文件。
每个 sheet 以标题行分隔，表格内容用「|」连接各列，空行跳过。
用法：python3 scripts/convert_xlsx.py
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
import config

try:
    import openpyxl
except ImportError:
    print("缺少依赖：pip3 install openpyxl")
    sys.exit(1)


def sheet_to_text(ws) -> str:
    lines = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        if not any(cells):
            continue
        lines.append(" | ".join(cells).rstrip(" |"))
    return "\n".join(lines)


def convert(xlsx_path: Path) -> Path:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    parts = []
    for name in wb.sheetnames:
        text = sheet_to_text(wb[name])
        if text.strip():
            parts.append(f"=== {name} ===\n{text}")
    out_path = xlsx_path.with_suffix(".txt")
    out_path.write_text("\n\n".join(parts), encoding="utf-8")
    return out_path


def main():
    data_dir = (ROOT / config.UPLOAD_DIR / "商务数据").resolve()
    if not data_dir.exists():
        print(f"目录不存在: {data_dir}")
        sys.exit(1)

    xlsx_files = sorted(data_dir.glob("*.xlsx"))
    if not xlsx_files:
        print("没有找到 .xlsx 文件")
        sys.exit(0)

    for xlsx_path in xlsx_files:
        out = convert(xlsx_path)
        print(f"  {xlsx_path.name} → {out.name}")

    print(f"\n完成：{len(xlsx_files)} 个文件已转换")


if __name__ == "__main__":
    main()
